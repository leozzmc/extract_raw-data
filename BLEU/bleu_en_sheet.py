from nltk.translate.bleu_score import *
from itertools import count
import os, re
import openpyxl
import pandas as pd



# 要加上修正後的sheet
ReferenceDataPath=[
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet1/sheet1.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet2/sheet2.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet3/sheet3.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet4/sheet4.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet5/sheet5.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet6/sheet6.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet7/sheet7.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet8/sheet8.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet9/sheet9.xlsx"
]

HypothesisDataPath=[
    "/azure_output_sheet/azure_output_1.xlsx",
    "/azure_output_sheet/azure_output_2.xlsx",
    "/azure_output_sheet/azure_output_3.xlsx",
    "/azure_output_sheet/azure_output_4.xlsx",
    "/azure_output_sheet/azure_output_5.xlsx",
    "/azure_output_sheet/azure_output_6.xlsx",
    "/azure_output_sheet/azure_output_7.xlsx",
    "/azure_output_sheet/azure_output_8.xlsx",
    "/azure_output_sheet/azure_output_9.xlsx",
]

class Excel_Data:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)

    def get_execel_file(self,DataPath):
        global sheet, ps
        # files = ROOTDIR +  "/../EN-Dictionary/DR詞彙解釋.xlsx"
        files = ROOTDIR +  DataPath
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        # Get first sheet
        sheet = ps[data.sheet_names[0]]

    # type:1 -> Reference, 2-> Hypothesis
    def ProcessSheet(self, DataPath, type:int):
        global OutputList
        OutputList= []
        self.get_execel_file(DataPath)
        # Reference
        if type == 1:
            for row in range(1,sheet.max_row+1):
                if sheet[row][1].value is not None:
                    txt = re.sub('[\.|,|\(|\)|\?|!|:|\'|“|”|—]','',sheet[row][1].value )
                    wordset= txt.split(' ')
                    OutputList.append(wordset)
            return OutputList
        # Hypothesis
        elif type==2:
            for row in range(1,sheet.max_row+1):
                if sheet[row][2].value is not None:
                    txt = re.sub('[\.|,|\(|\)|\?|!|:|\'|“|”|—]','',sheet[row][2].value )
                    wordset= txt.split(' ')
                    OutputList.append(wordset)
            return OutputList
        else:
            pass
    
    def Save_file(self,Path:str):
        ps.save(Path)
        

class Reference:

    def __init__(self):
        global data,reference
        data = Excel_Data()
        ## List for store "Return List"
        reference = []
        for i in ReferenceDataPath:
            reference.append(data.ProcessSheet(i,1))

    def check_reference(self):
        for i in range(0, len(reference)):
            print(f"\n------------------------------Reference{i}-------------------------------\n")
            for j in range(0,len(reference[i])):
                print(f"row{j} : {reference[i][j]}")


class Hypothesis:

    def __init__(self):
        global ex, hypothesis
        ex = Excel_Data()
        hypothesis = []
        for i in HypothesisDataPath:
            hypothesis.append(ex.ProcessSheet(i,2))
    
    def check_hypothesis(self):
        for i in range(0, len(hypothesis)):
            print(f"------------------------------Hypothesis{i}-------------------------------")
            for j in range(0,len(hypothesis[i])):
                print(f"row{j} : {hypothesis[i][j]}")



if __name__ == '__main__':
    ## initialize Reference and Hypothesis objects
    ref = Reference()
    ref.check_reference()
    hyp = Hypothesis()
    hyp.check_hypothesis()
    ex = Excel_Data()
    smo = SmoothingFunction()
    counter = 0

    #print(sentence_bleu([reference[0][1]], hypothesis[0][1], smoothing_function=smo.method5)*100)
    print(f"\n\n")
    for files in HypothesisDataPath:
        ex.get_execel_file(files)
        print(f"\ncounter: {counter}\n")
        for row in range(1,sheet.max_row+1):
           sheet[row][3].value = sentence_bleu([reference[counter][row-1]], hypothesis[counter][row-1], smoothing_function=smo.method5)*100

        ps.save(f"azure_output_{counter+1}.xlsx")
        counter +=1
        
    