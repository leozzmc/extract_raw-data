from nltk.translate.bleu_score import *
from itertools import count
import os, re
import openpyxl
import pandas as pd


ReferenceDataPath=[
    "/../EN-Dictionary/DR詞彙解釋.xlsx",
    "/../EN-Dictionary/GP詞彙解釋.xlsx",
    "/../EN-Dictionary/HL詞彙解釋.xlsx",
    "/../EN-Dictionary/ML詞彙表.xlsx",
    "/../EN-Dictionary/MP辭彙說明.xlsx",
    "/../EN-Dictionary/T詞彙解釋.xlsx",
    "/../EN-Dictionary/TGSM引用文典對照表.xlsx",
    "/../EN-Dictionary/TGSM詞彙選列對照表.xlsx",
    "/../EN-Dictionary/二十一度母 _ 八救難母 _ 詞彙表.xlsx"
]

HypothesisDataPath=[
    "/azure_output_1.xlsx",
    "/azure_output_2.xlsx",
    "/azure_output_3.xlsx",
    "/azure_output_4.xlsx",
    "/azure_output_5.xlsx",
    "/azure_output_6.xlsx",
    "/azure_output_7.xlsx",
    "/azure_output_8.xlsx",
    "/azure_output_9.xlsx",
]

class Excel_Data:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)

    def get_execel_file(self,DataPath):
        global sheet, output
        # files = ROOTDIR +  "/../EN-Dictionary/DR詞彙解釋.xlsx"
        files = ROOTDIR +  DataPath
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        # Get first sheet
        sheet = ps[data.sheet_names[0]]

    # type:1 -> Reference, 2-> Hypothesis
    def ProcessSheet(self, DataPath, type:int):
        global OutputList
        OutputList= []
        self.get_execel_file(DataPath)
        # Reference
        if type == 1:
            for row in range(2,sheet.max_row):
                if sheet[row][1].value is not None:
                    OutputList.append(str(sheet[row][1].value))
            return OutputList
        # Hypothesis
        elif type==2:
            for row in range(2,sheet.max_row):
                if sheet[row][1].value is not None:
                    OutputList.append(str(sheet[row][2].value))
            return OutputList
        else:
            pass
        

class Reference:

    def __init__(self):
        global data,reference
        data = Excel_Data()
        ## List for store "Return List"
        reference = []
        for i in ReferenceDataPath:
            reference.append(data.ProcessSheet(i,1))

    def check_reference(self):
        for i in range(0, len(reference)):
            print(f"------------------------------Reference{i}-------------------------------")
            for j in range(0,len(reference[i])):
                print(f"row{j} : {reference[i][j]}")

class Hypothesis:

    def __init__(self):
        global ex, hypothesis
        ex = Excel_Data()
        hypothesis = []
        for i in HypothesisDataPath:
            hypothesis.append(ex.ProcessSheet(i,2))
    
    def check_hypothesis(self):
        for i in range(0, len(hypothesis)):
            print(f"------------------------------Hypothesis{i}-------------------------------")
            for j in range(0,len(hypothesis[i])):
                print(f"row{j} : {hypothesis[i][j]}")



if __name__ == '__main__':
    ## initialize Reference and Hypothesis objects
    ref = Reference()
    ref.check_reference()
    hyp = Hypothesis()
    hyp.check_hypothesis()
    smo = SmoothingFunction()
    print(sentence_bleu([reference[0], reference[1], reference[2], reference[3], reference[4], reference[5], reference[6], reference[7], reference[8]], hypothesis[0], smoothing_function=smo.method5)*100)
    #print(sentence_bleu([reference[0], reference[1], reference[2], reference[3], reference[4], reference[5]],reference[6], smoothing_function=smo.method1))

    
# Smoothing Function method1: the lowest score, method5: the highest score