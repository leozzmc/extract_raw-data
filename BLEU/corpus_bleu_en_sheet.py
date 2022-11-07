from nltk.translate.bleu_score import *
from itertools import count
import os, re
import openpyxl
import pandas as pd



# 要加上修正後的sheet
ReferenceDataPath=[
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet1/sheet1.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet2/sheet2.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet3/sheet3.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet4/sheet4.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet5/sheet5.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet6/sheet6.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet7/sheet7.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet8/sheet8.xlsx",
    "/../EN-Dictionary/Test_for_BLEU/Glossary_Counsels/sheet9/sheet9.xlsx"
]

HypothesisDataPath=[
    "/azure_output_sheet/azure_output_1.xlsx",
    "/azure_output_sheet/azure_output_2.xlsx",
    "/azure_output_sheet/azure_output_3.xlsx",
    "/azure_output_sheet/azure_output_4.xlsx",
    "/azure_output_sheet/azure_output_5.xlsx",
    "/azure_output_sheet/azure_output_6.xlsx",
    "/azure_output_sheet/azure_output_7.xlsx",
    "/azure_output_sheet/azure_output_8.xlsx",
    "/azure_output_sheet/azure_output_9.xlsx",
]

class Excel_Data:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)

    def get_execel_file(self,DataPath):
        global sheet, ps
        # files = ROOTDIR +  "/../EN-Dictionary/DR詞彙解釋.xlsx"
        files = ROOTDIR +  DataPath
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        # Get first sheet
        sheet = ps[data.sheet_names[0]]

    # type:1 -> Reference, 2-> Hypothesis
    def ProcessSheet(self, DataPath, type:int):
        global OutputList
        OutputList= []
        self.get_execel_file(DataPath)
        # Reference
        if type == 1:
            for row in range(1,sheet.max_row+1):
                if sheet[row][1].value is not None:
                    txt = re.sub('[\.|,|\(|\)|\?|!|:|\'|“|”|—]','',sheet[row][1].value )
                    wordset= txt.split(' ')
                    OutputList.append(wordset)
            return OutputList
        # Hypothesis
        elif type==2:
            for row in range(1,sheet.max_row+1):
                if sheet[row][2].value is not None:
                    txt = re.sub('[\.|,|\(|\)|\?|!|:|\'|“|”|—]','',sheet[row][2].value )
                    wordset= txt.split(' ')
                    OutputList.append(wordset)
            return OutputList
        else:
            pass
    
    def Save_file(self,Path:str):
        ps.save(Path)
        

class Reference:

    def __init__(self):
        global data,reference
        data = Excel_Data()
        ## List for store "Return List"
        reference = []
        for i in ReferenceDataPath:
            reference.append(data.ProcessSheet(i,1))

    def check_reference(self):
        for i in range(0, len(reference)):
            print(f"\n------------------------------Reference{i}-------------------------------\n")
            for j in range(0,len(reference[i])):
                print(f"row{j} : {reference[i][j]}")


class Hypothesis:

    def __init__(self):
        global ex, hypothesis
        ex = Excel_Data()
        hypothesis = []
        for i in HypothesisDataPath:
            hypothesis.append(ex.ProcessSheet(i,2))
    
    def check_hypothesis(self):
        for i in range(0, len(hypothesis)):
            print(f"------------------------------Hypothesis{i}-------------------------------")
            for j in range(0,len(hypothesis[i])):
                print(f"row{j} : {hypothesis[i][j]}")



if __name__ == '__main__':
    ## initialize Reference and Hypothesis objects
    ref = Reference()
    hyp = Hypothesis()
    # ex = Excel_Data()
    smo = SmoothingFunction()

    
    ref0 = reference[0][0]
    ref1 = reference[0][1]
    ref2 = reference[0][2]
    ref3 = reference[0][3]
    hypo0 = hypothesis[0][0]
    hypo1 = hypothesis[0][1]
    hypo2 = hypothesis[0][2]
    hypo3 = hypothesis[0][3]
    #reflist = [[ref0],[ref1],[ref2],[ref3]]
    #hlist = [hypo0,hypo1,hypo2,hypo3]
    #print(corpus_bleu(reflist,hlist)*100)
    # reflist=[]
    # for i in range(0,len(reference)):  ##   1-8 
    #      reflist=[]
    #      for j in range(0,len(reference[i])): ## 8 3 60 34 17 40 21 26 7
    #         reflist.append([f"{reference[i][j]}"])
    #         print(corpus_bleu(reflist,hypothesis[i]))
    
    for i in range(0,len(reference)):
         for j in range(0,len(reference[i])):
            reference[i][j] = [reference[i][j]]
            
    for i in range(0,len(reference)):
        print(f"[SHEET-{i+1}] BLEU: {corpus_bleu(reference[i],hypothesis[i])*100} %")

    