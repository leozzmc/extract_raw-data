from itertools import count
import os, re
import openpyxl
import pandas as pd



class Dictionary1:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,DictionaryPath):
        global sheet, output
        files = ROOTDIR +  "/Dictionary/01-佛學辭典.xlsx"
        #files = ROOTDIR +  str(DictionaryPath)
        # Load orgin dictionary-1 excel file
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        # Save to another excel file
        ps.save('output_dic1.xlsx')
        output = openpyxl.load_workbook('output_dic1.xlsx')
        sheet = output[data.sheet_names[0]]
    
    # Bilingual Corpus, that is to merge tibetan columns,then output files for Azure.
    def MergeCols(self):
        print("-------------------Merge Columns --------------------------")
        for row in range(1,sheet.max_row):
            # The tibetan words and definitions are separated by Chinese symbol "："
            sheet[row][0].value = str(sheet[row][0].value) + "：" + str(sheet[row][1].value)
        sheet.delete_cols(2)
        output.save('output_dic1_corpus.xlsx')

    # Bilingual Dictionary, output files for aligment pipeline
    def ProcessDictionary(self):
        # Add Chinese entry columns
        #self.loadfile(DictionaryPath)
        print("Inserting cols .............")
        sheet.insert_cols(0,1)
        RowLimit = sheet.max_row
        for row in range(1,RowLimit):
            Chinese_words=""
            Substitude_word=""
            Count = 0
            # Chinese_wordsgroup = []
            for symbol in str(sheet[row][3].value):
                if Count == 0:
                    if symbol == "：":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words= ""
                        Substitude_word += "："
                        Count = 1
                    elif symbol == "。":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words = ""
                        Substitude_word += "。"
                        Count = 1 
                    else:
                        Chinese_words += symbol
                        Substitude_word +=symbol
             # Fix chinese difinition columns
            print(f"Substitude Word: {Substitude_word}")
            string = str(sheet[row][3].value)
            sheet[row][3].value = re.sub(f"{Substitude_word}","",string)
        
        # Fix Chinese entries
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Chinese_WordGroup = []
            Chinese_List = []  
            for col in range(0, sheet.max_column):
                for i in str(sheet[row][0].value):
                    if i == '，':
                        # print(str(sheet[row][0].value))
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Chinese_WordGroup = sheet[row][0].value.split("，")
                    # print(f"Row:[{row}] ->  {Chinese_WordGroup}")
                    # Store in tuples
                    for j in Chinese_WordGroup:
                        Chinese_List.append((j,sheet[row][1].value,sheet[row][2].value,sheet[row][3].value))
                    print(f"Chinese List: {Chinese_List}")
                    print("-------------------------")
                    # print(f"len: {len(Chinese_List)}")
                    termination_symbol_counter = 0
                     # Insert multiple rows before next row.
                    sheet.insert_rows(row,len(Chinese_List)+1)
                
                if len(Chinese_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Chinese_List)):
                        sheet[row + addrow][0].value = Chinese_List[addrow][0]
                        sheet[row + addrow][1].value = Chinese_List[addrow][1]
                        sheet[row + addrow][2].value = Chinese_List[addrow][2]
                        sheet[row + addrow][3].value = Chinese_List[addrow][3]
                    RowLimit = RowLimit + len(Chinese_List)
        output.save('output_dic1_dictionary.xlsx')

    # Iterate the sheet fields by rows and cols.
    # to sheet.max_row
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        for row in range(1,sheet.max_row):
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")
                # Delete Null items, that is delete fields without Tibetan definition.
                if sheet[row][1].value is None:
                    sheet.delete_rows(row)
                # Regular Expression for Chinese strings in the brackets
                string = str(sheet[row][2].value)
                # 全形括號
                if '（' in string: 
                    print(string)
                    p1 = re.compile(r'[（](.*?)[）]', re.S)
                    # find all match item and store in a list r1
                    r1 = re.findall(p1, string)
                    print(f"Match Results: {r1}")
                    # Process the list, delete numbers, such as '1895-1922', '1895一1922'...etc
                    for i in range(0,len(r1)):
                        new_item = re.sub(r'[0-9,"─","一","—"]+','',r1[i])
                        r1[i]= new_item          
                    print(f"New Results: {r1}")
                    # Delete Chinese strings in the brackets //刪除括號內中文
                    new_string = string
                    for i in r1:
                        new_string = re.sub(f"（{i}）","",new_string)                    
                    sheet[row][2].value = new_string
                    print(f"-------Results: {sheet[row][2].value}----")
        Answer = str(input("Choose Output, (1) Corpus (2) Dictionary, Please answer 1 or 2: "))
        if Answer == "1":
            self.MergeCols()
        elif Answer =="2":
            self.ProcessDictionary()
        else:
            print("Wrong Input. Program close.")


class Dictionary2:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)

    def get_execel_file(self,DictionaryPath):
        global sheet, output
        # /Dictionary/02-漢藏對照佛學辭典.xlsx
        #files = ROOTDIR +  str(DictionaryPath)
        files = ROOTDIR +  "/Dictionary/02-漢藏對照佛學辭典.xlsx"
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        ps.save('output_dic2.xlsx')
        output = openpyxl.load_workbook('output_dic2.xlsx')
        sheet = output[data.sheet_names[0]]
        
    
    # Iterate the sheet fields by rows and cols.
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        RowLimit = sheet.max_row
        sheet.delete_cols(3) 
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Tibetan_WordGroup = []
            Tibetan_List = []                    
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")                       
                # --------------(2) Tibetan words group seperations----------------------#
                # Check the numbers of termination symbol '།'
                for i in str(sheet[row][col].value):
                    if i == '།':
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Tibetan_WordGroup = sheet[row][1].value.split(" ")
                    print(f"Tibtean Group List: {Tibetan_WordGroup}")
                    # Store in tuples
                    for j in Tibetan_WordGroup:
                        # Store in tuples
                        Tibetan_List.append((sheet[row][0].value,j))
                    print(f"Tibetan List: {Tibetan_List}")
                    # Insert multiple rows before next row.
                    sheet.insert_rows(row+1,len(Tibetan_List))
                    termination_symbol_counter = 0
                
            
                if len(Tibetan_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Tibetan_List)):
                        sheet[row + addrow][0].value = Tibetan_List[addrow][0]
                        sheet[row + addrow][1].value = Tibetan_List[addrow][1]
                        #sheet[row + addrow][2].value = Tibetan_List[addrow][2]
                    RowLimit = RowLimit + len(Tibetan_List)

                # ------------------ (3) Delete words in the brackets-------------------#
                # string = str(sheet[row][2].value)
                # if ('（' or '〔' or '(' or '[' or '）' or '〕' or ')' or ']')in string:
                #     #print(string)
                #     p1 = re.compile(r'[（〔(\[](.*?)[）〕)\]]', re.S)
                #     r1 = re.findall(p1, string)
                #     print(f"Match Results: {r1}")
                #     # Delete Chinese strings in the brackets 
                #     new_string = string
                #     for i in r1:
                #         # new_string = re.sub(f"（{i}）|〔{i}〕|({i})|[{i}]","",new_string)
                #         new_string = re.sub(r"[（〔(\[](.*?)[）〕)\]]","",new_string)
                #         new_string = re.sub(r"[（〔(\[](.*?)[）〕)\]]","",new_string)             
                #     sheet[row][2].value = new_string
                #     print(f"-------Results: {sheet[row][2].value}----")
            # ----------------- (1) copy Chinese word-----------------------#
        for row in range(1,RowLimit): 
            if sheet[row][0].value is None:
                sheet[row][0].value = sheet[row-1][0].value 
        output.save('output_dic2.xlsx')


class Dictionary3:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,DictionaryPath):
        global sheet, output
        # /Dictionary/03-翻譯名義集 藏梵漢英.xlsx
        #files = ROOTDIR +  str(DictionaryPath)
        files = ROOTDIR +  "/Dictionary/03-翻譯名義集 藏梵漢英.xlsx"
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        ps.save('output_dic3.xlsx')
        output = openpyxl.load_workbook('output_dic3.xlsx')
        sheet = output[data.sheet_names[0]]
    
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        global RowLimit
        RowLimit = sheet.max_row
        # delete uneeded cols
        sheet.delete_cols(4)
        sheet.delete_cols(4)
        sheet.delete_cols(4)
        sheet.delete_cols(4)
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Chinese_WordGroup = []
            Chinese_List = []  
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------") 
                # --------------(1) Chinese words group seperations----------------------#
                # Check the numbers of termination symbol ','
                for i in str(sheet[row][2].value):
                    if i == '，':
                        print(str(sheet[row][2].value))
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Chinese_WordGroup = sheet[row][2].value.split("，")
                    print(f"Chinese Group List: {Chinese_WordGroup}")
                    # Store in tuples
                    for j in Chinese_WordGroup:
                        # Store in tuples
                        Chinese_List.append((sheet[row][0].value,sheet[row][1].value,j))
                    print(f"Chinese List: {Chinese_List}")
                    # Insert multiple rows before next row.
                    sheet.insert_rows(row,len(Chinese_List)+1)
                    termination_symbol_counter = 0
                
               
                if len(Chinese_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Chinese_List)):
                        sheet[row + addrow][0].value = Chinese_List[addrow][0]
                        sheet[row + addrow][1].value = Chinese_List[addrow][1]
                        sheet[row + addrow][2].value = Chinese_List[addrow][2]
                    RowLimit = RowLimit + len(Chinese_List)
        
        for row in range(1, sheet.max_row):
            for col in range(0,sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")      
                # ------------------ (3) delete words in the brackets -----------#
                string = str(sheet[row][col].value)
                if '(' in string:
                    print(string)
                    p1 = re.compile(r'[(](.*?)[)]', re.S)
                    r1 = re.findall(p1, string)
                    print(f"Match Results: {r1}")
                    # Delete Chinese strings in the brackets 
                    new_string = string
                    for i in r1:
                        new_string = re.sub(r"[(](.*?)[)]","",new_string)           
                    sheet[row][col].value = new_string
                    print(f"-------Results: {sheet[row][col].value}----")
        
        for row in range(1, sheet.max_row):
             # ----------------- (2) copy Tibetan word-----------------------#
            if sheet[row][0].value is None:
                sheet[row][0].value = sheet[row-1][0].value 
                  
        output.save('output_dic3.xlsx')
 


class Dictionary4:
    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,DictionaryPath):
        global sheet, output
        # /Dictionary/04-藏汉常用合稱辭典.xlsx
        #files = ROOTDIR +  str(DictionaryPath)
        files = ROOTDIR +  "/Dictionary/04-藏汉常用合稱辭典.xlsx"
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        ps.save('output_dic4.xlsx')
        output = openpyxl.load_workbook('output_dic4.xlsx')
        sheet = output[data.sheet_names[0]]

    def if_contain_Tibetan(self,strs):
        for _char in strs:
             if '\u0f00' <= _char <= '\u0fda':
                return True
        return False

    def delete_uneed_symbol(self,type: int):
        for row in range(1,sheet.max_row):
            Condition = self.if_contain_Tibetan(str(sheet[row][type].value))
            if Condition is True:
                # for output Dictionary4 -> sheet[row][2]
                # for output Dictionary4 -> sheet[row][1]
                if "參" in str(sheet[row][type].value) or "同" in str(sheet[row][type].value):
                    print(f"uneed symbol: {sheet[row][type].value}")
                    rule = re.compile(r'["同","參","和"][\u0F00-\u0fda].*$|["同","參","和"][ ][\u0F00-\u0fda].*$', re.S)
                    result = re.findall(rule, str(sheet[row][type].value))
                    print(f"Find Results: {''.join(result)}")
                    sheet[row][type].value = re.sub(f"{''.join(result)}","",sheet[row][type].value)
    
    # Bilingual Corpus, that is to merge tibetan columns,then output files for Azure.
    def MergeCols(self):
        print("-------------------Merge Columns --------------------------")
        for row in range(1,sheet.max_row):
            # The tibetan words and definitions are separated by Chinese symbol "："
            sheet[row][0].value = str(sheet[row][0].value) + "：" + str(sheet[row][1].value)
        sheet.delete_cols(2)
        # parameter:1
        self.delete_uneed_symbol(1)
        output.save('output_dic4_corpus.xlsx')

    # Bilingual Dictionary, output files for aligment pipeline
    def ProcessDictionary(self):
        # Add Chinese entry columns
        sheet.insert_cols(0,1)
        RowLimit = sheet.max_row
        for row in range(1,sheet.max_row):
            Chinese_words=""
            Substitude_word=""
            Count = 0
            # Chinese_wordsgroup = []
            for symbol in str(sheet[row][3].value):
                if Count == 0:
                    if symbol == "：":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words= ""
                        Substitude_word += "："
                        Count = 1
                    elif symbol == "。":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words = ""
                        Substitude_word += "。"
                        Count = 1 
                    else:
                        Chinese_words += symbol
                        Substitude_word +=symbol
             # Fix chinese difinition columns
            string = str(sheet[row][3].value)
            sheet[row][3].value = re.sub(f"{Substitude_word}","",string)
        
        # Fix Chinese entries
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Chinese_WordGroup = []
            Chinese_List = []  
            for col in range(0, sheet.max_column):
                for i in str(sheet[row][0].value):
                    if i == '，':
                        print(str(sheet[row][0].value))
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Chinese_WordGroup = sheet[row][0].value.split("，")
                    print(f"Row:[{row}] ->  {Chinese_WordGroup}")
                    # Store in tuples
                    for j in Chinese_WordGroup:
                        Chinese_List.append((j,sheet[row][1].value,sheet[row][2].value,sheet[row][3].value))
                    print(f"Chinese List: {Chinese_List}")
                    print("-------------------------")
                    print(f"len: {len(Chinese_List)}")
                    termination_symbol_counter = 0
                     # Insert multiple rows before next row.
                    sheet.insert_rows(row,len(Chinese_List)+1)
                
                if len(Chinese_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Chinese_List)):
                        sheet[row + addrow][0].value = Chinese_List[addrow][0]
                        sheet[row + addrow][1].value = Chinese_List[addrow][1]
                        sheet[row + addrow][2].value = Chinese_List[addrow][2]
                        sheet[row + addrow][3].value = Chinese_List[addrow][3]
                    RowLimit = RowLimit + len(Chinese_List)
        # parameter:2
        self.delete_uneed_symbol(3)
        output.save('output_dic4_dictionary.xlsx')
    
   
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        RowLimit = sheet.max_row
        for row in range(1,RowLimit):
            # ----------------- (1) copy Tibetan word-----------------------#
            if sheet[row][0].value is None:
                sheet[row][0].value = sheet[row-1][0].value 
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")
                 # ------------------ (2) delete words in the brackets -----------#
                string = str(sheet[row][col].value)
                if ('(' in string) or ('（' in string) :
                    print(string)
                    p1 = re.compile(r'[(（](.*?)[)）]', re.S)
                    r1 = re.findall(p1, string)
                    print(f"Match Results: {r1}")
                    # Delete Chinese strings in the brackets 
                    new_string = string
                    for i in r1:
                        new_string = re.sub(r"[（(](.*?)[)）]","",new_string)           
                    sheet[row][col].value = new_string
                    print(f"-------Results: {sheet[row][col].value}----")
  
        Answer = str(input("Choose Output, (1) Corpus (2) Dictionary, Please answer 1 or 2: "))
        if Answer == "1":
            self.MergeCols()
        elif Answer =="2":
            self.ProcessDictionary()
        else:
            print("Wrong Input. Program close.")

    

class Dictionary5:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,DictionaryPath):
        global sheet, output
        # /Dictionary/05-藏漢大詞典.xlsx
        #files = ROOTDIR +  str(DictionaryPath)
        files = ROOTDIR +  "/Dictionary/05-藏漢大詞典.xlsx"
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        ps.save('output_dic5.xlsx')
        output = openpyxl.load_workbook('output_dic5.xlsx')
        sheet = output[data.sheet_names[0]]
    
    # Bilingual Corpus, that is to merge tibetan columns,then output files for Azure.
    def MergeCols(self):
        print("-------------------Merge Columns --------------------------")
        for row in range(1,sheet.max_row):
            # The tibetan words and definitions are separated by Chinese symbol "："
            sheet[row][0].value = str(sheet[row][0].value) + "：" + str(sheet[row][1].value)
        sheet.delete_cols(2)
        output.save('output_dic5_corpus.xlsx')
    def loadfile(self,DictionaryPath):
        global sheet, output
        files = ROOTDIR +  "/output_dic5.xlsx"
        data = pd.ExcelFile(files)
        output = openpyxl.load_workbook(files)
        sheet = output[data.sheet_names[0]]

    # Bilingual Dictionary, output files for aligment pipeline
    def ProcessDictionary(self,DictionaryPath):
        #Add Chinese entry columns
        self.loadfile(DictionaryPath)
        print("Inserting cols .............")
        sheet.insert_cols(0,1)
        RowLimit = sheet.max_row
        for row in range(1,RowLimit):
            Chinese_words=""
            Substitude_word=""
            Count = 0
            # Chinese_wordsgroup = []x
            if ('《' in  str(sheet[row][3].value)) or ('》' in  str(sheet[row][3].value)):
                sheet[row][0].value = sheet[row][3].value
            elif ('：' in  str(sheet[row][3].value)) or ('。' in  str(sheet[row][3].value)):
                for symbol in str(sheet[row][3].value):
                    if Count == 0:
                        if symbol == "：":
                            # Chinese_wordsgroup.append(Chinese_words)
                            sheet[row][0].value = Chinese_words
                            Chinese_words= ""
                            Substitude_word += "："
                            Count = 1
                        elif symbol == "。":
                            # Chinese_wordsgroup.append(Chinese_words)
                            sheet[row][0].value = Chinese_words
                            Chinese_words = ""
                            Substitude_word += "。"
                            Count = 1 
                        else:
                            Chinese_words += symbol
                            Substitude_word +=symbol
                # Fix chinese difinition columns
                print(f"Substitude Word: {Substitude_word}")
                string = str(sheet[row][3].value)
                sheet[row][3].value = re.sub(f"{Substitude_word}","",string)
        
        # Fix Chinese entries
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Chinese_WordGroup = []
            Chinese_List = []  
            for col in range(0, sheet.max_column):
                for i in str(sheet[row][0].value):
                    if i == '，':
                        # print(str(sheet[row][0].value))
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Chinese_WordGroup = sheet[row][0].value.split("，")
                    # print(f"Row:[{row}] ->  {Chinese_WordGroup}")
                    # Store in tuples
                    for j in Chinese_WordGroup:
                        Chinese_List.append((j,sheet[row][1].value,sheet[row][2].value,sheet[row][3].value))
                    print(f"Chinese List: {Chinese_List}")
                    print("-------------------------")
                    # print(f"len: {len(Chinese_List)}")
                    termination_symbol_counter = 0
                     # Insert multiple rows before next row.
                    #sheet.insert_rows(row,len(Chinese_List)+1)
                    sheet.insert_rows(row,len(Chinese_List)+1)
                
                if len(Chinese_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Chinese_List)):
                        sheet[row + addrow][0].value = Chinese_List[addrow][0]
                        sheet[row + addrow][1].value = Chinese_List[addrow][1]
                        sheet[row + addrow][2].value = Chinese_List[addrow][2]
                        sheet[row + addrow][3].value = Chinese_List[addrow][3]
                    RowLimit = RowLimit + len(Chinese_List)
        output.save('output_dic5_dictionary.xlsx')
   
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        RowLimit = sheet.max_row
        for row in range(1,RowLimit):
            # ----------------- (1) copy Tibetan word-----------------------#
            if sheet[row][0].value is None:
                sheet[row][0].value = sheet[row-1][0].value 
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")
                 # ------------------ (2) delete words in the brackets -----------#
                string = str(sheet[row][col].value)
                if ('(' in string) or ('（' in string) :
                    print(string)
                    p1 = re.compile(r'[(（](.*?)[)）]', re.S)
                    r1 = re.findall(p1, string)
                    print(f"Match Results: {r1}")
                    # Delete Chinese strings in the brackets 
                    new_string = string
                    for i in r1:
                        new_string = re.sub(r"[（(](.*?)[)）]","",new_string)           
                    sheet[row][col].value = new_string
                    print(f"-------Results: {sheet[row][col].value}----")
                else:
                    print(f"-------Results: {sheet[row][col].value}----")
        # Answer = str(input("Choose Output, (1) Corpus (2) Dictionary, Please answer 1 or 2: "))
        # if Answer == "1":
        #     self.MergeCols()
        # elif Answer =="2":
        #     self.ProcessDictionary()
        # else:
        #     print("Wrong Input. Program close.")
        output.save('output_dic5.xlsx')
        print("Output phase1 file.")
        self.ProcessDictionary()


class Dictionary6:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,DictionaryPath):
        global sheet, output
        # /Dictionary/06-藏漢佛學辭典.xlsx
        #files = ROOTDIR +  str(DictionaryPath)
        files = ROOTDIR +  "/Dictionary/06-藏漢佛學辭典.xlsx"
        data = pd.ExcelFile(files)
        ps = openpyxl.load_workbook(files)
        ps.save('output_dic6.xlsx')
        output = openpyxl.load_workbook('output_dic6.xlsx')
        sheet = output[data.sheet_names[0]]
    
    # Bilingual Corpus, that is to merge tibetan columns,then output files for Azure.
    def MergeCols(self):
        print("-------------------Merge Columns --------------------------")
        for row in range(1,New_Row):
            # The tibetan words and definitions are separated by Chinese symbol "："
            sheet[row][0].value = str(sheet[row][0].value) + "：" + str(sheet[row][1].value)
        sheet.delete_cols(2)
        output.save('output_dic6_corpus.xlsx')
    # def loadfile(self,DictionaryPath):
    #     global sheet, output
    #     files = ROOTDIR +  "/output_dic6.xlsx"
    #     data = pd.ExcelFile(files)
    #     output = openpyxl.load_workbook(files)
    #     sheet = output[data.sheet_names[0]]

    # Bilingual Dictionary, output files for aligment pipeline
    def ProcessDictionary(self):
        # Add Chinese entry columns
        #self.loadfile(DictionaryPath)
        print("Inserting cols .............")
        sheet.insert_cols(0,1)
        RowLimit = sheet.max_row
        global New_RowLimit
        New_RowLimit = sheet.max_row
        for row in range(1,RowLimit):
            Chinese_words=""
            Substitude_word=""
            Count = 0
            # Chinese_wordsgroup = []
            for symbol in str(sheet[row][3].value):
                if Count == 0:
                    if symbol == "：":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words= ""
                        Substitude_word += "："
                        Count = 1
                    elif symbol == "。":
                        # Chinese_wordsgroup.append(Chinese_words)
                        sheet[row][0].value = Chinese_words
                        Chinese_words = ""
                        Substitude_word += "。"
                        Count = 1 
                    else:
                        Chinese_words += symbol
                        Substitude_word +=symbol
             # Fix chinese difinition columns
            print(f"Substitude Word: {Substitude_word}")
            string = str(sheet[row][3].value)
            sheet[row][3].value = re.sub(f"{Substitude_word}","",string)
        
        # Fix Chinese entries
        for row in range(1,RowLimit):
            termination_symbol_counter=0
            Chinese_WordGroup = []
            Chinese_List = []  
            for col in range(0, sheet.max_column):
                for i in str(sheet[row][0].value):
                    if i == '，':
                        # print(str(sheet[row][0].value))
                        termination_symbol_counter = termination_symbol_counter +1
                # Split the words in a group
                if termination_symbol_counter > 1:
                    Chinese_WordGroup = sheet[row][0].value.split("，")
                    # print(f"Row:[{row}] ->  {Chinese_WordGroup}")
                    # Store in tuples
                    for j in Chinese_WordGroup:
                        Chinese_List.append((j,sheet[row][1].value,sheet[row][2].value,sheet[row][3].value))
                    print(f"Chinese List: {Chinese_List}")
                    print("-------------------------")
                    # print(f"len: {len(Chinese_List)}")
                    termination_symbol_counter = 0
                     # Insert multiple rows before next row.
                    sheet.insert_rows(row,len(Chinese_List)+1)
                
                if len(Chinese_List) > 0:
                    sheet.delete_rows(row)
                    for addrow in range(0,len(Chinese_List)):
                        sheet[row + addrow][0].value = Chinese_List[addrow][0]
                        sheet[row + addrow][1].value = Chinese_List[addrow][1]
                        sheet[row + addrow][2].value = Chinese_List[addrow][2]
                        sheet[row + addrow][3].value = Chinese_List[addrow][3]
                    #RowLimit = RowLimit + len(Chinese_List)
                    New_RowLimit =  New_RowLimit + len(Chinese_List)
        output.save('output_dic6_dictionary.xlsx')
    

   
    def ProcessSheet(self,DictionaryPath):
        self.get_execel_file(DictionaryPath)
        RowLimit = sheet.max_row
        global New_Row
        New_Row = sheet.max_row
        for row in range(1,RowLimit):
            #-------------------(0) delete empty Chinese explaination ------------#
            if sheet[row][2].value is None:
                    sheet.delete_rows(row)
                    New_Row = New_Row -1
            # ----------------- (1) copy Tibetan word-----------------------#
            if sheet[row][0].value is None and sheet[row][1].value is not None:
                sheet[row][0].value = sheet[row-1][0].value 
            for col in range(0, sheet.max_column):
                print(f"---------------------------------Row:{row} Col:{col}------------------------------------")
                 # ------------------ (2) delete words in the brackets -----------#
                string = str(sheet[row][col].value)
                if ('(' in string) or ('（' in string) :
                    print(string)
                    p1 = re.compile(r'[(（](.*?)[)）]', re.S)
                    r1 = re.findall(p1, string)
                    print(f"Match Results: {r1}")
                    # Delete Chinese strings in the brackets 
                    new_string = string
                    for i in r1:
                        new_string = re.sub(r"[（(](.*?)[)）]","",new_string)           
                    sheet[row][col].value = new_string
                    print(f"-------Results: {sheet[row][col].value}----")
        Answer = str(input("Choose Output, (1) Corpus (2) Dictionary, Please answer 1 or 2: "))
        if Answer == "1":
            self.MergeCols()
        elif Answer =="2":
            self.ProcessDictionary()
        else:
            print("Wrong Input. Program close.")
        # output.save('output_dic5.xlsx')
        # print("Output phase1 file.")
        # self.ProcessDictionary()


if __name__ == '__main__':
    a= int(input("(1)(2)(3)(4)(5)(6) Which dictionaries? Type number: "))
    if a == 1:
        Dic1 = Dictionary1()
        DicPath = str(input("Enter Dictionary File Path: "))
        Dic1.ProcessSheet(DicPath)
    elif a==2:
        Dic2 = Dictionary2()
        DicPath = str(input("Enter Dictionary File Path: "))
        Dic2.ProcessSheet(DicPath)
    elif a==3:
        Dic3 = Dictionary3()
        DicPath = str(input("Enter Dictionary File Path: "))
        Dic3.ProcessSheet(DicPath)
    elif a==4:
        Dic4 = Dictionary4()
        DicPath = str(input("Enter Dictionary File Path: "))
        Dic4.ProcessSheet(DicPath)
    elif a==5:
        Dic5 = Dictionary5()
        DicPath = str(input("Enter Dictionary File Path: "))
        #Dic5.ProcessSheet(DicPath)
        Dic5.ProcessDictionary(DicPath)
    elif a==6:
        Dic6 = Dictionary6()
        DicPath = str(input("Enter Dictionary File Path: "))
        Dic6.ProcessSheet(DicPath)
    else:
        print("No such dictionaries")