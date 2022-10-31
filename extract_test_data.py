from itertools import count
import os, re
import openpyxl
import pandas as pd


class Sheet:

    def __init__(self):
        global ROOTDIR 
        ROOTDIR = os.getcwd()
        print(ROOTDIR)
    
    def get_execel_file(self,sheet_Name):
        global sheet, output
        if sheet_Name == 1:
            files = ROOTDIR +  "/EN-Dictionary/Test_for_BLEU/Glossary_Counsels.xlsx"
            data = pd.ExcelFile(files)
            ps = openpyxl.load_workbook(files)
            ps.save('output_sheet.xlsx')
            output = openpyxl.load_workbook('output_sheet.xlsx')   
            sheet = output[data.sheet_names[sheet_Name]]        ## sheet1
        else:
            data = pd.ExcelFile('output_sheet.xlsx')
            output = openpyxl.load_workbook('output_sheet.xlsx')   
            sheet = output[data.sheet_names[sheet_Name]]


    def ProcessSheet(self,sheet_Name):
        self.get_execel_file(sheet_Name)

        for row in range(1,sheet.max_row):
            print(f"---------------------------------Row:{row}------------------------------------")
            string = str(sheet[row][0].value)
            # 全形括號
            if '（' in string: 
                print(string)
                p1 = re.compile(r'[（](.*?)[）]', re.S)
                # find all match item and store in a list r1
                r1 = re.findall(p1, string)
                print(f"Match Results: {r1}")
                # Delete Chinese strings in the brackets //刪除括號內中文
                new_string = string
                for i in r1:
                    new_string = re.sub(f"（{i}）","",new_string)                    
                sheet[row][0].value = new_string
                print(f"-------Results: {sheet[row][0].value}----")
        output.save('output_sheet.xlsx')

if __name__ == '__main__':
    Sh = Sheet()
    for i in range(1,10):
        SheetParam = i
        Sh.ProcessSheet(int(SheetParam))