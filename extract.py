from importlib.metadata import files
from re import I
import openpyxl
import os
import pandas as pd 
import click


@click.group()
@click.option('--target', help="Sutras name you want to extract e.g. --target T12")
def cli(target):
    global files, sutras
    global targetPath, outputPath
    global RootDIR
    RootDIR = os.getcwd()
    if target == 'T12':
        sutras = 'T12'
        files=[]
        for i in range(1,33):   
            files.append(f"八千頌般若經藏漢對照第{i}品.xlsx")
        targetPath = os.chdir(RootDIR+ "/tm-master/tm-master/sources/T12")
        outputPath = RootDIR + "/output/T12/"

    elif target == 'T4090':
        files=[]
        sutras = 'T4090'
        files.append("Vol.1 俱舍論第一品藏漢對照(1).xlsx")
        files.append("Vol.2 俱舍論第二品藏漢對照(1).xlsx")
        files.append("Vol.3 俱舍論第三品藏漢對照(1).xlsx")
        files.append("Vol.4 俱舍論第四品藏漢對照(1).xlsx")
        files.append("Vol.5 俱舍論第五品藏漢對照(1).xlsx")
        files.append("Vol.6 俱舍論第六品藏漢對照(1).xlsx")
        files.append("Vol.7 俱舍論第七品藏漢對照(1).xlsx")
        files.append("Vol.8 俱舍論第八品藏漢對照(1).xlsx")
        files.append("Vol.9 俱舍論第九品藏漢對照(1).xlsx")
        targetPath = os.chdir("tm-master/tm-master/sources/T4090")
        outputPath = RootDIR + "/output" + "/T4090/"
    print(files)
    return files, targetPath, outputPath, RootDIR



@cli.command()
def extract():
    '''To extract data form target excel files '''
    # RootDIR = os.getcwd()
    # os.chdir("tm-master/tm-master/sources/T12")
    # outputPath = RootDIR + "/output/"
    outputName = "vol"

    # for i in range(1,33):
    #     files.append(f"八千頌般若經藏漢對照第{i}品.xlsx")

    # length in the range() function is determine by the number of volumes you want to output
    for i in range(0,len(files)):
        data = pd.ExcelFile(files[i])
        ps = openpyxl.load_workbook(files[i])
        # select sheets in excel files
        sheet = ps[data.sheet_names[0]]
        # Check if the output dir. is already exist
        isExists = os.path.exists(outputPath+outputName+str(i+1))
        # print(sheet.max_row)

        if not isExists:
            # Make Directories for each volumes
            OUPUT_DIRNAME=outputPath+outputName+str(i+1)
            os.makedirs(OUPUT_DIRNAME)
            # Output files
            with open(f"{OUPUT_DIRNAME}/{outputName}{i+1}.txt", 'w', encoding='utf16') as f:
                # Iterate rows in the sheet
                for row in range(2, sheet.max_row + 1):
                    # Diferent procedure for different sutras
                    if sutras == "T12":
                        # Exceptions cases for vol 30~32, there are missing chinese translation in the column C in these files.
                        if i > 28:
                            # ignore the None type table
                            if sheet['B' + str(row)].value is not None:
                                # Get one of Chinese tranlation of each row in the sheet
                                sutras_row = sheet['B' + str(row)].value
                                f.write(sutras_row +"\r\n")
                        else:
                            # ignore the None type table
                            if sheet['C' + str(row)].value is not None:
                                # Get one of Chinese tranlation of each row in the sheet
                                sutras_row = sheet['C' + str(row)].value
                                f.write(sutras_row +"\r\n")
                    elif sutras == 'T4090':
                        # ignore the None type table
                            if sheet['B' + str(row)].value is not None:
                                # Get one of Chinese tranlation of each row in the sheet
                                sutras_row = sheet['B' + str(row)].value
                                f.write(sutras_row +"\r\n")
           

if __name__ == '__main__':
    cli()


    # Test Output
    # for row in range(2, sheet.max_row + 1):
    #             if sheet['C' + str(row)].value is not None:
    #                 sutras_row = sheet['C' + str(row)].value
    #                 print(f"-[Vol: {i+1}][Row: {row}]---------------------------------------------------------")
    #                 print(sutras_row)

    


            
   


    
    


